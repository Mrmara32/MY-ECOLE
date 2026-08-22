import os
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, g

from database import db, gen_id, rows_to_list, row_to_dict, log_action, get_settings
from auth import require_auth, require_role

bp = Blueprint('finances_routes', __name__, url_prefix='/api')

FIN_ROLES = ('admin', 'directeur', 'comptable', 'secretaire')
# Opérations bancaires directes (créer une dépense/recette, encaisser un paiement) :
# réservées au comptable, avec l'administrateur en accès total. Le directeur conserve
# un rôle de contrôle (consultation, approbation/rejet) mais n'exécute plus lui-même
# d'opération financière — une séparation des tâches demandée explicitement.
FIN_CREATE_ROLES = ('admin', 'comptable')


def determiner_statut_validation(type_, montant, role_createur):
    """Applique les règles d'approbation comptable (points 3/4/5 du cahier des charges):
    - L'admin (fondateur) a toujours autorité totale : ses opérations sont auto-validées.
    - Les recettes (entrée) ne nécessitent pas d'approbation.
    - Dépenses (sortie) :
        < seuil_directeur              -> auto-validé
        seuil_directeur..seuil_admin    -> attente_directeur (sauf si le créateur EST le directeur -> auto)
        > seuil_admin                   -> attente_admin (sauf si le créateur EST l'admin -> auto, déjà couvert ci-dessus)
    """
    if role_createur == 'admin':
        return 'auto'
    if type_ == 'entree':
        return 'auto'
    s = get_settings()
    seuil_directeur = float(s.get('seuil_approbation_directeur') or 30000)
    seuil_admin = float(s.get('seuil_approbation_admin') or 100000)
    if montant < seuil_directeur:
        return 'auto'
    if montant <= seuil_admin:
        return 'auto' if role_createur == 'directeur' else 'attente_directeur'
    return 'attente_admin'


# ─────────────────────────────────────────────────────────────
# TRANSACTIONS
# ─────────────────────────────────────────────────────────────
def generer_transactions_recurrentes_dues(ecole_id):
    """Vérifie les transactions récurrentes actives (loyer, salaires fixes...) et crée
    automatiquement celle du mois en cours si elle n'a pas déjà été générée, dès que le
    jour du mois configuré est atteint. Appelée à chaque consultation du journal comptable
    (pas de tâche planifiée disponible sur cet hébergement) — sans effet si rien n'est dû."""
    aujourdhui = datetime.now()
    mois_courant = aujourdhui.strftime('%Y-%m')
    jour_actuel = aujourdhui.day

    dues = db.execute(
        """SELECT * FROM transactions_recurrentes
           WHERE ecole_id=? AND actif=1 AND jour_du_mois<=? AND (dernier_mois_genere IS NULL OR dernier_mois_genere!=?)""",
        (ecole_id, jour_actuel, mois_courant)
    ).fetchall()
    for tr in dues:
        date_op = f"{mois_courant}-{str(tr['jour_du_mois']).zfill(2)}"
        tid = gen_id('t')
        statut = determiner_statut_validation(tr['type'], tr['montant'], 'admin')
        db.execute(
            """INSERT INTO transactions (id,ecole_id,type,date_op,description,categorie,moyen_paiement,montant,statut_validation,cree_par)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (tid, ecole_id, tr['type'], date_op, (tr['description'] or '') + ' (généré automatiquement)',
             tr['categorie'], tr['moyen_paiement'], tr['montant'], statut, tr['cree_par']),
        )
        db.execute("UPDATE transactions_recurrentes SET dernier_mois_genere=? WHERE id=?", (mois_courant, tr['id']))
    if dues:
        db.commit()


@bp.route('/transactions', methods=['GET'])
@require_auth
def list_transactions():
    generer_transactions_recurrentes_dues(g.user['ecole_id'])
    type_ = request.args.get('type')
    categorie = request.args.get('categorie')
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    q = request.args.get('q')
    statut_validation = request.args.get('statut_validation')

    sql = ("SELECT t.*, e.nom as eleve_nom, e.prenom as eleve_prenom, "
           "u1.full_name as cree_par_nom, u2.full_name as valide_par_nom "
           "FROM transactions t LEFT JOIN eleves e ON e.id=t.eleve_id "
           "LEFT JOIN users u1 ON u1.id=t.cree_par LEFT JOIN users u2 ON u2.id=t.valide_par WHERE t.ecole_id=?")
    params = [g.user['ecole_id']]
    if type_: sql += " AND t.type=?"; params.append(type_)
    if categorie: sql += " AND t.categorie=?"; params.append(categorie)
    if date_debut: sql += " AND t.date_op>=?"; params.append(date_debut)
    if date_fin: sql += " AND t.date_op<=?"; params.append(date_fin)
    if q: sql += " AND (t.description LIKE ? OR t.reference LIKE ?)"; params += [f"%{q}%", f"%{q}%"]
    if statut_validation: sql += " AND t.statut_validation=?"; params.append(statut_validation)
    sql += " ORDER BY t.date_op DESC, t.created_at DESC"

    rows = db.execute(sql, params).fetchall()
    return jsonify(rows_to_list(rows))


@bp.route('/transactions-recurrentes', methods=['GET'])
@require_auth
@require_role('admin', 'directeur', 'comptable')
def list_transactions_recurrentes():
    rows = db.execute("SELECT * FROM transactions_recurrentes WHERE ecole_id=? ORDER BY actif DESC, categorie", (g.user['ecole_id'],)).fetchall()
    return jsonify(rows_to_list(rows))


@bp.route('/transactions-recurrentes', methods=['POST'])
@require_auth
@require_role(*FIN_CREATE_ROLES)
def create_transaction_recurrente():
    body = request.get_json(silent=True) or {}
    type_, categorie = body.get('type'), body.get('categorie')
    try:
        montant = float(body.get('montant'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Montant invalide'}), 400
    if type_ not in ('entree', 'sortie') or not categorie or montant <= 0:
        return jsonify({'error': 'Champs requis invalides'}), 400
    jour = int(body.get('jour_du_mois') or 1)
    tid = gen_id('trec')
    db.execute(
        """INSERT INTO transactions_recurrentes (id,ecole_id,type,categorie,description,montant,moyen_paiement,jour_du_mois,cree_par)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (tid, g.user['ecole_id'], type_, categorie, body.get('description'), montant, body.get('moyen_paiement'), jour, g.user['id']),
    )
    db.commit()
    row = db.execute("SELECT * FROM transactions_recurrentes WHERE id=?", (tid,)).fetchone()
    return jsonify(row_to_dict(row)), 201


@bp.route('/transactions-recurrentes/<tr_id>', methods=['PUT'])
@require_auth
@require_role(*FIN_CREATE_ROLES)
def update_transaction_recurrente(tr_id):
    body = request.get_json(silent=True) or {}
    db.execute(
        """UPDATE transactions_recurrentes SET
           categorie=COALESCE(?,categorie), description=COALESCE(?,description),
           montant=COALESCE(?,montant), moyen_paiement=COALESCE(?,moyen_paiement),
           jour_du_mois=COALESCE(?,jour_du_mois),
           actif=COALESCE(?,actif) WHERE id=? AND ecole_id=?""",
        (body.get('categorie'), body.get('description'), body.get('montant'), body.get('moyen_paiement'),
         body.get('jour_du_mois'), (1 if body.get('actif') else 0) if 'actif' in body else None, tr_id, g.user['ecole_id']),
    )
    db.commit()
    row = db.execute("SELECT * FROM transactions_recurrentes WHERE id=? AND ecole_id=?", (tr_id, g.user['ecole_id'])).fetchone()
    if not row:
        return jsonify({'error': 'Introuvable'}), 404
    return jsonify(row_to_dict(row))


@bp.route('/transactions-recurrentes/<tr_id>', methods=['DELETE'])
@require_auth
@require_role('admin')
def delete_transaction_recurrente(tr_id):
    db.execute("DELETE FROM transactions_recurrentes WHERE id=? AND ecole_id=?", (tr_id, g.user['ecole_id']))
    db.commit()
    return jsonify({'success': True})


@bp.route('/transactions/export', methods=['GET'])
@require_auth
@require_role('admin', 'directeur', 'comptable')
def export_transactions_excel():
    """Export du journal comptable au format Excel, pour l'expert-comptable
    ou un usage hors-ligne. Reprend les mêmes filtres que la liste."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    type_ = request.args.get('type')
    categorie = request.args.get('categorie')
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')

    sql = ("SELECT t.*, e.nom as eleve_nom, e.prenom as eleve_prenom, u1.full_name as cree_par_nom "
           "FROM transactions t LEFT JOIN eleves e ON e.id=t.eleve_id "
           "LEFT JOIN users u1 ON u1.id=t.cree_par WHERE t.ecole_id=? AND t.statut_validation IN ('auto','valide')")
    params = [g.user['ecole_id']]
    if type_: sql += " AND t.type=?"; params.append(type_)
    if categorie: sql += " AND t.categorie=?"; params.append(categorie)
    if date_debut: sql += " AND t.date_op>=?"; params.append(date_debut)
    if date_fin: sql += " AND t.date_op<=?"; params.append(date_fin)
    sql += " ORDER BY t.date_op ASC, t.created_at ASC"
    lignes = db.execute(sql, params).fetchall()

    s = get_settings(g.user['ecole_id'])
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal comptable"

    NAVY = "0E2A5E"
    entete_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    entete_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    titre_font = Font(name="Arial", bold=True, size=14, color=NAVY)
    bordure = Border(bottom=Side(style="thin", color="D1D5DB"))

    ws["A1"] = s.get("ecole_nom") or "Groupe Scolaire"
    ws["A1"].font = titre_font
    ws["A2"] = "Journal comptable"
    ws["A2"].font = Font(name="Arial", size=11, color="6B7280")
    periode = f"Du {date_debut} au {date_fin}" if date_debut or date_fin else "Toutes périodes"
    ws["A3"] = periode
    ws["A3"].font = Font(name="Arial", size=10, italic=True, color="6B7280")

    entetes = ["Date", "Type", "Catégorie", "Description", "Élève lié", "Moyen de paiement",
               "Référence", "Montant entrée (GNF)", "Montant sortie (GNF)", "Enregistré par"]
    ligne_entete = 5
    for col, texte in enumerate(entetes, start=1):
        c = ws.cell(row=ligne_entete, column=col, value=texte)
        c.font = entete_font
        c.fill = entete_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = ligne_entete + 1
    premiere_ligne_data = r
    for t in lignes:
        eleve = f"{t['eleve_prenom']} {t['eleve_nom']}" if t['eleve_nom'] else ""
        ws.cell(row=r, column=1, value=t["date_op"])
        ws.cell(row=r, column=2, value="Recette" if t["type"] == "entree" else "Dépense")
        ws.cell(row=r, column=3, value=t["categorie"] or "")
        ws.cell(row=r, column=4, value=t["description"] or "")
        ws.cell(row=r, column=5, value=eleve)
        ws.cell(row=r, column=6, value=t["moyen_paiement"] or "")
        ws.cell(row=r, column=7, value=t["reference"] or "")
        ws.cell(row=r, column=8, value=t["montant"] if t["type"] == "entree" else None)
        ws.cell(row=r, column=9, value=t["montant"] if t["type"] == "sortie" else None)
        ws.cell(row=r, column=10, value=t["cree_par_nom"] or "")
        for col in (8, 9):
            cel = ws.cell(row=r, column=col)
            cel.number_format = '#,##0 "GNF"'
        for col in range(1, 11):
            ws.cell(row=r, column=col).border = bordure
        r += 1
    derniere_ligne_data = r - 1

    r += 1
    if derniere_ligne_data >= premiere_ligne_data:
        ws.cell(row=r, column=7, value="TOTAUX").font = Font(name="Arial", bold=True)
        c8 = ws.cell(row=r, column=8, value=f"=SUM(H{premiere_ligne_data}:H{derniere_ligne_data})")
        c9 = ws.cell(row=r, column=9, value=f"=SUM(I{premiere_ligne_data}:I{derniere_ligne_data})")
        c8.number_format = c9.number_format = '#,##0 "GNF"'
        c8.font = c9.font = Font(name="Arial", bold=True)
        r += 1
        ws.cell(row=r, column=7, value="SOLDE").font = Font(name="Arial", bold=True)
        c10 = ws.cell(row=r, column=8, value=f"=H{r-1}-I{r-1}")
        c10.number_format = '#,##0 "GNF"'
        c10.font = Font(name="Arial", bold=True, color=NAVY)

    largeurs = [12, 10, 20, 30, 20, 16, 14, 18, 18, 18]
    for i, l in enumerate(largeurs, start=1):
        ws.column_dimensions[chr(64 + i)].width = l

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_action(g.user, 'export', 'journal_comptable', '', {'periode': periode})
    nom_fichier = f"Journal-Comptable-{date_debut or 'debut'}_{date_fin or 'fin'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nom_fichier,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/transactions/en-attente', methods=['GET'])
@require_auth
@require_role('admin', 'directeur')
def transactions_en_attente():
    """Liste les transactions en attente d'approbation, filtrées selon le rôle de qui consulte."""
    if g.user['role'] == 'admin':
        sql = ("SELECT t.*, e.nom as eleve_nom, e.prenom as eleve_prenom, u.full_name as cree_par_nom "
               "FROM transactions t LEFT JOIN eleves e ON e.id=t.eleve_id LEFT JOIN users u ON u.id=t.cree_par "
               "WHERE t.ecole_id=? AND t.statut_validation IN ('attente_directeur','attente_admin') ORDER BY t.created_at DESC")
        rows = db.execute(sql, (g.user['ecole_id'],)).fetchall()
    else:  # directeur : ne voit que ce qu'il doit lui-même approuver
        sql = ("SELECT t.*, e.nom as eleve_nom, e.prenom as eleve_prenom, u.full_name as cree_par_nom "
               "FROM transactions t LEFT JOIN eleves e ON e.id=t.eleve_id LEFT JOIN users u ON u.id=t.cree_par "
               "WHERE t.ecole_id=? AND t.statut_validation='attente_directeur' ORDER BY t.created_at DESC")
        rows = db.execute(sql, (g.user['ecole_id'],)).fetchall()
    return jsonify(rows_to_list(rows))


@bp.route('/transactions', methods=['POST'])
@require_auth
@require_role(*FIN_CREATE_ROLES)
def create_transaction():
    body = request.get_json(silent=True) or {}
    type_, date_op = body.get('type'), body.get('date_op')
    try:
        montant = float(body.get('montant')) if body.get('montant') is not None else None
    except (TypeError, ValueError):
        return jsonify({'error': 'Montant invalide'}), 400
    if not type_ or not date_op or not montant:
        return jsonify({'error': 'Type, date et montant requis'}), 400
    if montant <= 0:
        return jsonify({'error': 'Le montant doit être positif'}), 400

    statut = determiner_statut_validation(type_, montant, g.user['role'])
    tid = gen_id('t')
    db.execute(
        "INSERT INTO transactions (id,ecole_id,type,date_op,description,categorie,moyen_paiement,montant,reference,eleve_id,cree_par,statut_validation) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        (tid, g.user['ecole_id'], type_, date_op, body.get('description'), body.get('categorie'), body.get('moyen_paiement'),
         montant, body.get('reference'), body.get('eleve_id'), g.user['id'], statut),
    )
    db.commit()
    log_action(g.user, 'creation', 'transaction', tid, {'type': type_, 'montant': montant, 'statut_validation': statut})
    row = db.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
    result = row_to_dict(row)
    if statut in ('attente_directeur', 'attente_admin'):
        result['_info'] = ("Cette dépense dépasse le seuil autorisé et doit être approuvée par "
                            + ("l'administrateur" if statut == 'attente_admin' else "le directeur")
                            + " avant d'être comptabilisée dans les totaux.")
    return jsonify(result), 201


@bp.route('/transactions/<t_id>/approuver', methods=['PUT'])
@require_auth
@require_role('admin', 'directeur')
def approuver_transaction(t_id):
    t = db.execute("SELECT * FROM transactions WHERE id=? AND ecole_id=?", (t_id, g.user['ecole_id'])).fetchone()
    if not t:
        return jsonify({'error': 'Introuvable'}), 404
    if t['statut_validation'] == 'attente_directeur' and g.user['role'] not in ('directeur', 'admin'):
        return jsonify({'error': 'Accès refusé'}), 403
    if t['statut_validation'] == 'attente_admin' and g.user['role'] != 'admin':
        return jsonify({'error': "Seul l'administrateur peut approuver cette opération"}), 403
    if t['statut_validation'] not in ('attente_directeur', 'attente_admin'):
        return jsonify({'error': "Cette opération n'est pas en attente d'approbation"}), 400

    db.execute("UPDATE transactions SET statut_validation='valide', valide_par=?, date_validation=CURRENT_TIMESTAMP WHERE id=?",
               (g.user['id'], t_id))
    db.commit()
    log_action(g.user, 'approbation', 'transaction', t_id, {'montant': t['montant'], 'description': t['description']})
    row = db.execute("SELECT * FROM transactions WHERE id=?", (t_id,)).fetchone()
    return jsonify(row_to_dict(row))


@bp.route('/transactions/<t_id>/rejeter', methods=['PUT'])
@require_auth
@require_role('admin', 'directeur')
def rejeter_transaction(t_id):
    body = request.get_json(silent=True) or {}
    t = db.execute("SELECT * FROM transactions WHERE id=? AND ecole_id=?", (t_id, g.user['ecole_id'])).fetchone()
    if not t:
        return jsonify({'error': 'Introuvable'}), 404
    if t['statut_validation'] == 'attente_admin' and g.user['role'] != 'admin':
        return jsonify({'error': "Seul l'administrateur peut rejeter cette opération"}), 403
    if t['statut_validation'] not in ('attente_directeur', 'attente_admin'):
        return jsonify({'error': "Cette opération n'est pas en attente d'approbation"}), 400

    db.execute("UPDATE transactions SET statut_validation='rejete', valide_par=?, date_validation=CURRENT_TIMESTAMP, motif_rejet=? WHERE id=?",
               (g.user['id'], body.get('motif'), t_id))
    db.commit()
    log_action(g.user, 'rejet', 'transaction', t_id, {'motif': body.get('motif'), 'montant': t['montant']})
    row = db.execute("SELECT * FROM transactions WHERE id=?", (t_id,)).fetchone()
    return jsonify(row_to_dict(row))


@bp.route('/transactions/<t_id>', methods=['PUT'])
@require_auth
@require_role('admin')
def update_transaction(t_id):
    """Toute modification d'une transaction existante est réservée à l'administrateur (point 3)."""
    body = request.get_json(silent=True) or {}
    existing = db.execute("SELECT * FROM transactions WHERE id=? AND ecole_id=?", (t_id, g.user['ecole_id'])).fetchone()
    if not existing:
        return jsonify({'error': 'Introuvable'}), 404
    db.execute(
        "UPDATE transactions SET description=COALESCE(?,description), categorie=COALESCE(?,categorie), "
        "moyen_paiement=COALESCE(?,moyen_paiement), montant=COALESCE(?,montant), reference=COALESCE(?,reference), "
        "date_op=COALESCE(?,date_op) WHERE id=? AND ecole_id=?",
        (body.get('description'), body.get('categorie'), body.get('moyen_paiement'),
         body.get('montant'), body.get('reference'), body.get('date_op'), t_id, g.user['ecole_id']),
    )
    db.commit()
    log_action(g.user, 'modification', 'transaction', t_id, {'avant': dict(existing), 'apres': body})
    row = db.execute("SELECT * FROM transactions WHERE id=?", (t_id,)).fetchone()
    return jsonify(row_to_dict(row))


@bp.route('/transactions/<t_id>', methods=['DELETE'])
@require_auth
@require_role('admin')
def delete_transaction(t_id):
    existing = db.execute("SELECT * FROM transactions WHERE id=? AND ecole_id=?", (t_id, g.user['ecole_id'])).fetchone()
    if not existing:
        return jsonify({'error': 'Introuvable'}), 404
    db.execute("DELETE FROM transactions WHERE id=? AND ecole_id=?", (t_id, g.user['ecole_id']))
    db.commit()
    log_action(g.user, 'suppression', 'transaction', t_id, dict(existing))
    return jsonify({'success': True})


# ─────────────────────────────────────────────────────────────
# BARÈMES FRAIS
# ─────────────────────────────────────────────────────────────
@bp.route('/frais', methods=['GET'])
@require_auth
def list_frais():
    rows = db.execute("SELECT * FROM frais_scolarite WHERE ecole_id=? ORDER BY annee_scolaire DESC, classe", (g.user['ecole_id'],)).fetchall()
    return jsonify(rows_to_list(rows))


@bp.route('/frais', methods=['POST'])
@require_auth
@require_role('admin', 'directeur', 'comptable')
def create_frais():
    body = request.get_json(silent=True) or {}
    classe, annee_scolaire = body.get('classe'), body.get('annee_scolaire')
    if not classe or not annee_scolaire:
        return jsonify({'error': 'Classe et année requis'}), 400
    fid = gen_id('fs')
    try:
        db.execute(
            "INSERT INTO frais_scolarite (id,ecole_id,classe,annee_scolaire,frais_inscription,scolarite_annuelle,nombre_tranches) "
            "VALUES (?,?,?,?,?,?,?)",
            (fid, g.user['ecole_id'], classe, annee_scolaire, body.get('frais_inscription', 0),
             body.get('scolarite_annuelle', 0), body.get('nombre_tranches', 3)),
        )
        db.commit()
    except Exception:
        return jsonify({'error': 'Barème déjà existant pour cette classe/année'}), 409
    row = db.execute("SELECT * FROM frais_scolarite WHERE id=?", (fid,)).fetchone()
    return jsonify(row_to_dict(row)), 201


@bp.route('/frais/<f_id>', methods=['PUT'])
@require_auth
@require_role('admin')
def update_frais(f_id):
    body = request.get_json(silent=True) or {}
    db.execute(
        "UPDATE frais_scolarite SET frais_inscription=COALESCE(?,frais_inscription), "
        "scolarite_annuelle=COALESCE(?,scolarite_annuelle), nombre_tranches=COALESCE(?,nombre_tranches) WHERE id=? AND ecole_id=?",
        (body.get('frais_inscription'), body.get('scolarite_annuelle'), body.get('nombre_tranches'), f_id, g.user['ecole_id']),
    )
    db.commit()
    row = db.execute("SELECT * FROM frais_scolarite WHERE id=? AND ecole_id=?", (f_id, g.user['ecole_id'])).fetchone()
    if not row:
        return jsonify({'error': 'Introuvable'}), 404
    return jsonify(row_to_dict(row))


@bp.route('/frais/<f_id>', methods=['DELETE'])
@require_auth
@require_role('admin')
def delete_frais(f_id):
    db.execute("DELETE FROM frais_scolarite WHERE id=? AND ecole_id=?", (f_id, g.user['ecole_id']))
    db.commit()
    return jsonify({'success': True})


# ─────────────────────────────────────────────────────────────
# PAIEMENTS ÉCHELONNÉS
# ─────────────────────────────────────────────────────────────
@bp.route('/paiements', methods=['GET'])
@require_auth
def list_paiements():
    eleve_id = request.args.get('eleve_id')
    classe = request.args.get('classe')
    statut = request.args.get('statut')
    annee_scolaire = request.args.get('annee_scolaire')

    sql = """SELECT p.*, e.nom, e.prenom, e.classe, e.matricule
             FROM paiements p JOIN eleves e ON e.id=p.eleve_id WHERE p.ecole_id=?"""
    params = [g.user['ecole_id']]
    if eleve_id: sql += " AND p.eleve_id=?"; params.append(eleve_id)
    if classe: sql += " AND e.classe=?"; params.append(classe)
    if statut: sql += " AND p.statut=?"; params.append(statut)
    if annee_scolaire: sql += " AND p.annee_scolaire=?"; params.append(annee_scolaire)
    sql += " ORDER BY p.date_echeance ASC"

    rows = db.execute(sql, params).fetchall()
    return jsonify(rows_to_list(rows))


@bp.route('/paiements/generer', methods=['POST'])
@require_auth
@require_role(*FIN_ROLES)
def generer_paiements():
    body = request.get_json(silent=True) or {}
    eleve_id, annee_scolaire = body.get('eleve_id'), body.get('annee_scolaire')
    if not eleve_id or not annee_scolaire:
        return jsonify({'error': 'Élève et année requis'}), 400

    eleve = db.execute("SELECT * FROM eleves WHERE id=? AND ecole_id=?", (eleve_id, g.user['ecole_id'])).fetchone()
    if not eleve:
        return jsonify({'error': 'Élève introuvable'}), 404

    bareme = db.execute(
        "SELECT * FROM frais_scolarite WHERE ecole_id=? AND classe=? AND annee_scolaire=?", (g.user['ecole_id'], eleve['classe'], annee_scolaire)
    ).fetchone()
    if not bareme:
        return jsonify({'error': f"Aucun barème pour la classe {eleve['classe']} en {annee_scolaire}"}), 404

    existing = db.execute(
        "SELECT COUNT(*) as c FROM paiements WHERE eleve_id=? AND annee_scolaire=?", (eleve_id, annee_scolaire)
    ).fetchone()['c']
    if existing > 0:
        return jsonify({'error': 'Paiements déjà générés pour cet élève/année'}), 409

    date_debut = body.get('date_debut')
    start = datetime.fromisoformat(date_debut) if date_debut else datetime.now()
    count = 0

    if bareme['frais_inscription'] > 0:
        pid = gen_id('pai')
        db.execute(
            "INSERT INTO paiements (id,ecole_id,eleve_id,annee_scolaire,type_frais,libelle,montant_du,date_echeance) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (pid, g.user['ecole_id'], eleve_id, annee_scolaire, 'inscription', "Frais d'inscription",
             bareme['frais_inscription'], start.strftime('%Y-%m-%d')),
        )
        count += 1

    # Répartition des tranches de scolarité : 45% / 40% / 15% du montant annuel
    # (remplace l'ancienne répartition en parts égales)
    POURCENTAGES_TRANCHES = [0.45, 0.40, 0.15]
    nb = len(POURCENTAGES_TRANCHES)
    for i, pct in enumerate(POURCENTAGES_TRANCHES):
        d = start + timedelta(days=i * 91)  # approx. tous les 3 mois
        montant_tranche = round(bareme['scolarite_annuelle'] * pct)
        pid = gen_id('pai') + str(i)
        db.execute(
            "INSERT INTO paiements (id,ecole_id,eleve_id,annee_scolaire,type_frais,libelle,montant_du,date_echeance) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (pid, g.user['ecole_id'], eleve_id, annee_scolaire, f'tranche_{i+1}', f'Tranche {i+1} de scolarité ({int(pct*100)}%)',
             montant_tranche, d.strftime('%Y-%m-%d')),
        )
        count += 1

    db.commit()
    return jsonify({'success': True, 'count': count}), 201


@bp.route('/paiements', methods=['POST'])
@require_auth
@require_role(*FIN_ROLES)
def create_paiement():
    body = request.get_json(silent=True) or {}
    eleve_id, montant_du = body.get('eleve_id'), body.get('montant_du')
    if not eleve_id or not montant_du:
        return jsonify({'error': 'Élève et montant requis'}), 400
    pid = gen_id('pai')
    db.execute(
        "INSERT INTO paiements (id,ecole_id,eleve_id,annee_scolaire,type_frais,libelle,montant_du,date_echeance) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (pid, g.user['ecole_id'], eleve_id, body.get('annee_scolaire'), body.get('type_frais', 'autre'),
         body.get('libelle', 'Paiement'), montant_du, body.get('date_echeance')),
    )
    db.commit()
    row = db.execute("SELECT * FROM paiements WHERE id=?", (pid,)).fetchone()
    return jsonify(row_to_dict(row)), 201


@bp.route('/paiements/<p_id>', methods=['DELETE'])
@require_auth
@require_role('admin')
def delete_paiement(p_id):
    db.execute("DELETE FROM paiements WHERE id=? AND ecole_id=?", (p_id, g.user['ecole_id']))
    db.commit()
    return jsonify({'success': True})


@bp.route('/paiements/<p_id>/verser', methods=['POST'])
@require_auth
@require_role(*FIN_CREATE_ROLES)
def verser_paiement(p_id):
    body = request.get_json(silent=True) or {}
    try:
        montant = float(body.get('montant')) if body.get('montant') is not None else None
    except (TypeError, ValueError):
        return jsonify({'error': 'Montant invalide'}), 400
    if not montant or montant <= 0:
        return jsonify({'error': 'Montant invalide'}), 400

    pai = db.execute("SELECT * FROM paiements WHERE id=? AND ecole_id=?", (p_id, g.user['ecole_id'])).fetchone()
    if not pai:
        return jsonify({'error': 'Paiement introuvable'}), 404

    # Le montant en Franc Guinéen n'a pas de sous-unité utile en pratique : on arrondit
    # systématiquement pour ne jamais laisser un résidu de calcul flottant (ex: 0,33 GNF)
    # bloquer indéfiniment un paiement par ailleurs déjà réglé. Aucune limite maximale
    # n'est imposée : un léger dépassement (arrondi, pourboire, avance) reste acceptable.
    montant = round(montant)

    date_vers = body.get('date_vers') or datetime.now().strftime('%Y-%m-%d')
    moyen_paiement = body.get('moyen_paiement', 'Espèces')

    vid = gen_id('v')
    db.execute(
        "INSERT INTO versements (id,ecole_id,paiement_id,eleve_id,date_vers,montant,moyen_paiement,reference,recu_par) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        (vid, g.user['ecole_id'], p_id, pai['eleve_id'], date_vers, montant, moyen_paiement, body.get('reference'), g.user['id']),
    )
    new_paye = round(pai['montant_paye'] + montant)
    statut = 'paye' if new_paye >= round(pai['montant_du']) - 1 else ('partiel' if new_paye > 0 else 'a_payer')
    db.execute("UPDATE paiements SET montant_paye=?, statut=? WHERE id=?", (new_paye, statut, p_id))

    eleve = db.execute("SELECT nom,prenom,matricule FROM eleves WHERE id=?", (pai['eleve_id'],)).fetchone()
    tid = gen_id('t')
    db.execute(
        "INSERT INTO transactions (id,ecole_id,type,date_op,description,categorie,moyen_paiement,montant,reference,eleve_id) "
        "VALUES (?,?,?,?,?,?,?,?,?,?)",
        (tid, g.user['ecole_id'], 'entree', date_vers, f"{pai['libelle']} — {eleve['prenom']} {eleve['nom']} ({eleve['matricule']})",
         "Frais d'inscription" if pai['type_frais'] == 'inscription' else 'Frais de scolarité',
         moyen_paiement, montant, body.get('reference') or f"REC-{vid}", pai['eleve_id']),
    )
    db.commit()

    row = db.execute("SELECT * FROM paiements WHERE id=?", (p_id,)).fetchone()
    log_action(g.user, 'versement', 'paiement', p_id, {'montant': montant, 'eleve_id': pai['eleve_id'], 'moyen_paiement': moyen_paiement})
    return jsonify({'paiement': row_to_dict(row)})


@bp.route('/paiements/solde/<eleve_id>', methods=['GET'])
@require_auth
def solde_paiements(eleve_id):
    row = db.execute(
        """SELECT
           COALESCE(SUM(montant_du),0) as total_du,
           COALESCE(SUM(montant_paye),0) as total_paye,
           COALESCE(SUM(montant_du-montant_paye),0) as reste,
           SUM(CASE WHEN statut='paye' THEN 1 ELSE 0 END) as payees,
           SUM(CASE WHEN statut!='paye' THEN 1 ELSE 0 END) as en_attente
           FROM paiements WHERE eleve_id=? AND ecole_id=?""",
        (eleve_id, g.user['ecole_id']),
    ).fetchone()
    result = dict(row)
    for k in ('payees', 'en_attente'):
        result[k] = result[k] or 0
    return jsonify(result)


@bp.route('/paiements/soldes', methods=['GET'])
@require_auth
def soldes_tous():
    rows = db.execute(
        """SELECT e.id, e.matricule, e.nom, e.prenom, e.classe,
           COALESCE(SUM(p.montant_du),0) as total_du, COALESCE(SUM(p.montant_paye),0) as total_paye,
           COALESCE(SUM(p.montant_du-p.montant_paye),0) as reste
           FROM eleves e LEFT JOIN paiements p ON p.eleve_id=e.id
           WHERE e.ecole_id=?
           GROUP BY e.id ORDER BY reste DESC, e.nom""",
        (g.user['ecole_id'],)
    ).fetchall()
    return jsonify(rows_to_list(rows))


@bp.route('/versements/<eleve_id>', methods=['GET'])
@require_auth
def versements_eleve(eleve_id):
    rows = db.execute(
        """SELECT v.*, p.libelle, u.full_name as recu_par_nom
           FROM versements v LEFT JOIN paiements p ON p.id=v.paiement_id LEFT JOIN users u ON u.id=v.recu_par
           WHERE v.eleve_id=? AND v.ecole_id=? ORDER BY v.date_vers DESC""",
        (eleve_id, g.user['ecole_id']),
    ).fetchall()
    return jsonify(rows_to_list(rows))


# ─────────────────────────────────────────────────────────────
# CANTINE
# ─────────────────────────────────────────────────────────────
@bp.route('/cantine/menus', methods=['GET'])
@require_auth
def list_menus():
    debut, fin = request.args.get('debut'), request.args.get('fin')
    sql = "SELECT * FROM cantine_menus WHERE ecole_id=?"
    params = [g.user['ecole_id']]
    if debut: sql += " AND date_menu>=?"; params.append(debut)
    if fin: sql += " AND date_menu<=?"; params.append(fin)
    sql += " ORDER BY date_menu DESC"
    rows = db.execute(sql, params).fetchall()
    return jsonify(rows_to_list(rows))


@bp.route('/cantine/menus', methods=['POST'])
@require_auth
@require_role(*FIN_ROLES)
def create_menu():
    body = request.get_json(silent=True) or {}
    if not body.get('date_menu'):
        return jsonify({'error': 'Date requise'}), 400
    mid = gen_id('menu')
    try:
        db.execute(
            "INSERT INTO cantine_menus (id,ecole_id,date_menu,entree,plat,dessert) VALUES (?,?,?,?,?,?)",
            (mid, g.user['ecole_id'], body['date_menu'], body.get('entree'), body.get('plat'), body.get('dessert')),
        )
        db.commit()
    except Exception:
        return jsonify({'error': 'Menu existant pour cette date'}), 409
    row = db.execute("SELECT * FROM cantine_menus WHERE id=?", (mid,)).fetchone()
    return jsonify(row_to_dict(row)), 201


@bp.route('/cantine/menus/<m_id>', methods=['PUT'])
@require_auth
@require_role(*FIN_ROLES)
def update_menu(m_id):
    body = request.get_json(silent=True) or {}
    db.execute(
        "UPDATE cantine_menus SET entree=COALESCE(?,entree), plat=COALESCE(?,plat), dessert=COALESCE(?,dessert) WHERE id=? AND ecole_id=?",
        (body.get('entree'), body.get('plat'), body.get('dessert'), m_id, g.user['ecole_id']),
    )
    db.commit()
    row = db.execute("SELECT * FROM cantine_menus WHERE id=? AND ecole_id=?", (m_id, g.user['ecole_id'])).fetchone()
    if not row:
        return jsonify({'error': 'Introuvable'}), 404
    return jsonify(row_to_dict(row))


@bp.route('/cantine/menus/<m_id>', methods=['DELETE'])
@require_auth
@require_role(*FIN_ROLES)
def delete_menu(m_id):
    db.execute("DELETE FROM cantine_menus WHERE id=? AND ecole_id=?", (m_id, g.user['ecole_id']))
    db.commit()
    return jsonify({'success': True})


@bp.route('/cantine/abonnements', methods=['GET'])
@require_auth
def list_abonnements():
    mois = request.args.get('mois')
    classe = request.args.get('classe')
    paye = request.args.get('paye')

    sql = """SELECT c.*, e.nom, e.prenom, e.classe, e.matricule
             FROM cantine_abonnements c JOIN eleves e ON e.id=c.eleve_id WHERE c.ecole_id=?"""
    params = [g.user['ecole_id']]
    if mois: sql += " AND c.mois=?"; params.append(mois)
    if classe: sql += " AND e.classe=?"; params.append(classe)
    if paye not in (None, ''): sql += " AND c.paye=?"; params.append(1 if paye == '1' else 0)
    sql += " ORDER BY c.mois DESC, e.nom"

    rows = db.execute(sql, params).fetchall()
    return jsonify(rows_to_list(rows))


@bp.route('/cantine/abonnements', methods=['POST'])
@require_auth
@require_role(*FIN_ROLES)
def create_abonnement():
    body = request.get_json(silent=True) or {}
    eleve_id, mois = body.get('eleve_id'), body.get('mois')
    if not eleve_id or not mois:
        return jsonify({'error': 'Élève et mois requis'}), 400
    aid = gen_id('abo')
    try:
        db.execute(
            "INSERT INTO cantine_abonnements (id,ecole_id,eleve_id,mois,formule,montant) VALUES (?,?,?,?,?,?)",
            (aid, g.user['ecole_id'], eleve_id, mois, body.get('formule', 'complète'), body.get('montant', 0)),
        )
        db.commit()
    except Exception:
        return jsonify({'error': 'Abonnement déjà existant pour cet élève/mois'}), 409
    row = db.execute("SELECT * FROM cantine_abonnements WHERE id=?", (aid,)).fetchone()
    return jsonify(row_to_dict(row)), 201


@bp.route('/cantine/abonnements/<a_id>', methods=['PUT'])
@require_auth
@require_role(*FIN_ROLES)
def update_abonnement(a_id):
    body = request.get_json(silent=True) or {}
    db.execute(
        "UPDATE cantine_abonnements SET formule=COALESCE(?,formule), montant=COALESCE(?,montant), "
        "paye=COALESCE(?,paye) WHERE id=? AND ecole_id=?",
        (body.get('formule'), body.get('montant'), (1 if body.get('paye') else 0) if 'paye' in body else None, a_id, g.user['ecole_id']),
    )
    db.commit()
    row = db.execute("SELECT * FROM cantine_abonnements WHERE id=? AND ecole_id=?", (a_id, g.user['ecole_id'])).fetchone()
    if not row:
        return jsonify({'error': 'Introuvable'}), 404
    return jsonify(row_to_dict(row))


@bp.route('/cantine/abonnements/<a_id>/payer', methods=['POST'])
@require_auth
@require_role(*FIN_CREATE_ROLES)
def payer_abonnement(a_id):
    body = request.get_json(silent=True) or {}
    abo = db.execute(
        "SELECT c.*, e.nom, e.prenom, e.matricule FROM cantine_abonnements c "
        "JOIN eleves e ON e.id=c.eleve_id WHERE c.id=? AND c.ecole_id=?", (a_id, g.user['ecole_id'])
    ).fetchone()
    if not abo:
        return jsonify({'error': 'Introuvable'}), 404
    if abo['paye']:
        return jsonify({'error': 'Déjà payé'}), 400

    db.execute("UPDATE cantine_abonnements SET paye=1 WHERE id=? AND ecole_id=?", (a_id, g.user['ecole_id']))
    d = body.get('date_vers') or datetime.now().strftime('%Y-%m-%d')
    tid = gen_id('t')
    db.execute(
        "INSERT INTO transactions (id,ecole_id,type,date_op,description,categorie,moyen_paiement,montant,reference,eleve_id) "
        "VALUES (?,?,?,?,?,?,?,?,?,?)",
        (tid, g.user['ecole_id'], 'entree', d, f"Cantine {abo['mois']} — {abo['prenom']} {abo['nom']} ({abo['matricule']})",
         'Cantine', body.get('moyen_paiement', 'Espèces'), abo['montant'],
         body.get('reference') or f"CANT-{abo['id']}", abo['eleve_id']),
    )
    db.commit()
    return jsonify({'success': True})


@bp.route('/cantine/abonnements/<a_id>', methods=['DELETE'])
@require_auth
@require_role(*FIN_ROLES)
def delete_abonnement(a_id):
    db.execute("DELETE FROM cantine_abonnements WHERE id=? AND ecole_id=?", (a_id, g.user['ecole_id']))
    db.commit()
    return jsonify({'success': True})


# ─────────────────────────────────────────────────────────────
# BUDGET PRÉVISIONNEL
# ─────────────────────────────────────────────────────────────
@bp.route('/budgets', methods=['GET'])
@require_auth
@require_role('admin', 'directeur', 'comptable')
def list_budgets():
    mois = request.args.get('mois')
    sql = "SELECT * FROM budgets WHERE ecole_id=?"
    params = [g.user['ecole_id']]
    if mois: sql += " AND mois=?"; params.append(mois)
    sql += " ORDER BY type, categorie"
    rows = db.execute(sql, params).fetchall()
    return jsonify(rows_to_list(rows))


@bp.route('/budgets', methods=['POST'])
@require_auth
@require_role('admin', 'directeur', 'comptable')
def create_ou_maj_budget():
    """Crée ou met à jour (si déjà défini pour cette catégorie/mois) une ligne de budget."""
    body = request.get_json(silent=True) or {}
    categorie, type_, mois = body.get('categorie'), body.get('type'), body.get('mois')
    try:
        montant = float(body.get('montant_prevu'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Montant prévu invalide'}), 400
    if not categorie or type_ not in ('entree', 'sortie') or not mois or montant < 0:
        return jsonify({'error': 'Champs requis invalides'}), 400

    existant = db.execute(
        "SELECT id FROM budgets WHERE ecole_id=? AND categorie=? AND type=? AND mois=?", (g.user['ecole_id'], categorie, type_, mois)
    ).fetchone()
    if existant:
        db.execute("UPDATE budgets SET montant_prevu=? WHERE id=?", (montant, existant['id']))
        bid = existant['id']
    else:
        bid = gen_id('bud')
        db.execute(
            "INSERT INTO budgets (id,ecole_id,categorie,type,mois,montant_prevu) VALUES (?,?,?,?,?,?)",
            (bid, g.user['ecole_id'], categorie, type_, mois, montant),
        )
    db.commit()
    row = db.execute("SELECT * FROM budgets WHERE id=?", (bid,)).fetchone()
    return jsonify(row_to_dict(row)), 201


@bp.route('/budgets/<b_id>', methods=['DELETE'])
@require_auth
@require_role('admin', 'directeur', 'comptable')
def delete_budget(b_id):
    db.execute("DELETE FROM budgets WHERE id=? AND ecole_id=?", (b_id, g.user['ecole_id']))
    db.commit()
    return jsonify({'success': True})


@bp.route('/budgets/comparaison', methods=['GET'])
@require_auth
@require_role('admin', 'directeur', 'comptable')
def comparaison_budget():
    """Pour un mois donné : budget prévu vs réalisé (transactions effectivement
    comptabilisées) par catégorie, avec l'écart — le cœur du suivi budgétaire."""
    mois = request.args.get('mois') or datetime.now().strftime('%Y-%m')
    budgets = db.execute("SELECT * FROM budgets WHERE ecole_id=? AND mois=?", (g.user['ecole_id'], mois)).fetchall()
    reel = db.execute(
        """SELECT categorie, type, SUM(montant) as total FROM transactions
           WHERE ecole_id=? AND strftime('%Y-%m', date_op)=? AND statut_validation IN ('auto','valide')
           GROUP BY categorie, type""",
        (g.user['ecole_id'], mois)
    ).fetchall()
    reel_par_cle = {(r['categorie'], r['type']): r['total'] for r in reel}

    cles = set((b['categorie'], b['type']) for b in budgets) | set(reel_par_cle.keys())
    resultat = []
    budgets_par_cle = {(b['categorie'], b['type']): b['montant_prevu'] for b in budgets}
    for categorie, type_ in sorted(cles):
        prevu = budgets_par_cle.get((categorie, type_), 0)
        realise = reel_par_cle.get((categorie, type_), 0)
        resultat.append({
            'categorie': categorie, 'type': type_,
            'montant_prevu': prevu, 'montant_realise': realise,
            'ecart': realise - prevu,
            'pourcentage': round((realise / prevu) * 100, 1) if prevu else None,
        })
    return jsonify({'mois': mois, 'lignes': resultat})


@bp.route('/analyse-comptable', methods=['GET'])
@require_auth
@require_role('admin', 'directeur', 'comptable')
def analyse_comptable():
    """Données pour le tableau de bord comptable : tendance sur 12 mois et
    répartition des dépenses/recettes par catégorie sur la période choisie."""
    mois = request.args.get('mois') or datetime.now().strftime('%Y-%m')

    tendance = rows_to_list(db.execute(
        """SELECT strftime('%Y-%m',date_op) as mois,
           SUM(CASE WHEN type='entree' THEN montant ELSE 0 END) as recettes,
           SUM(CASE WHEN type='sortie' THEN montant ELSE 0 END) as depenses
           FROM transactions WHERE ecole_id=? AND date_op >= date('now','-12 months') AND statut_validation IN ('auto','valide')
           GROUP BY strftime('%Y-%m',date_op) ORDER BY mois""",
        (g.user['ecole_id'],)
    ).fetchall())

    repartition_depenses = rows_to_list(db.execute(
        """SELECT categorie, SUM(montant) as total FROM transactions
           WHERE ecole_id=? AND type='sortie' AND strftime('%Y-%m',date_op)=? AND statut_validation IN ('auto','valide')
           GROUP BY categorie ORDER BY total DESC""",
        (g.user['ecole_id'], mois)
    ).fetchall())

    repartition_recettes = rows_to_list(db.execute(
        """SELECT categorie, SUM(montant) as total FROM transactions
           WHERE ecole_id=? AND type='entree' AND strftime('%Y-%m',date_op)=? AND statut_validation IN ('auto','valide')
           GROUP BY categorie ORDER BY total DESC""",
        (g.user['ecole_id'], mois)
    ).fetchall())

    return jsonify({
        'mois': mois, 'tendance': tendance,
        'repartition_depenses': repartition_depenses,
        'repartition_recettes': repartition_recettes,
    })


# ─────────────────────────────────────────────────────────────
# RAPPROCHEMENT BANCAIRE
# ─────────────────────────────────────────────────────────────
@bp.route('/rapprochement/analyser', methods=['POST'])
@require_auth
@require_role('admin', 'comptable')
def analyser_rapprochement():
    """Reçoit un relevé bancaire (CSV ou Excel — 3 colonnes : Date, Description, Montant ;
    montant positif = crédit/entrée, négatif = débit/sortie) et propose un rapprochement
    automatique avec les opérations déjà enregistrées et non encore rapprochées."""
    import pandas as pd
    import io

    fichier = request.files.get('fichier')
    if not fichier or fichier.filename == '':
        return jsonify({'error': 'Aucun fichier reçu'}), 400
    ext = os.path.splitext(fichier.filename)[1].lower()
    if ext not in ('.csv', '.xlsx', '.xls'):
        return jsonify({'error': 'Format non supporté (CSV ou Excel attendu)'}), 400

    try:
        contenu = fichier.read()
        if ext == '.csv':
            df = pd.read_csv(io.BytesIO(contenu), sep=None, engine='python')
        else:
            df = pd.read_excel(io.BytesIO(contenu))
    except Exception as e:
        return jsonify({'error': f"Impossible de lire le fichier : {e}"}), 400

    if df.shape[1] < 3:
        return jsonify({'error': "Le fichier doit contenir au moins 3 colonnes : Date, Description, Montant"}), 400
    df = df.iloc[:, :3]
    df.columns = ['date', 'description', 'montant']

    lignes_releve = []
    for _, row in df.iterrows():
        try:
            date_op = pd.to_datetime(row['date']).strftime('%Y-%m-%d')
            montant = float(str(row['montant']).replace(' ', '').replace(',', '.'))
        except (ValueError, TypeError):
            continue
        if montant == 0:
            continue
        lignes_releve.append({
            'date': date_op, 'description': str(row['description'] or ''),
            'montant': abs(montant), 'type': 'entree' if montant > 0 else 'sortie',
        })

    non_rapprochees = rows_to_list(db.execute(
        """SELECT * FROM transactions WHERE ecole_id=? AND (rapproche IS NULL OR rapproche=0)
           AND statut_validation IN ('auto','valide') ORDER BY date_op""",
        (g.user['ecole_id'],)
    ).fetchall())

    resultats = []
    ids_deja_suggeres = set()
    for ligne in lignes_releve:
        candidats = [
            t for t in non_rapprochees
            if t['id'] not in ids_deja_suggeres
            and t['type'] == ligne['type']
            and abs(t['montant'] - ligne['montant']) < 1
            and abs((datetime.strptime(t['date_op'], '%Y-%m-%d') - datetime.strptime(ligne['date'], '%Y-%m-%d')).days) <= 5
        ]
        if candidats:
            meilleur = min(candidats, key=lambda t: abs((datetime.strptime(t['date_op'], '%Y-%m-%d') - datetime.strptime(ligne['date'], '%Y-%m-%d')).days))
            ids_deja_suggeres.add(meilleur['id'])
            resultats.append({'releve': ligne, 'transaction_suggeree': meilleur, 'statut': 'correspondance_trouvee'})
        else:
            resultats.append({'releve': ligne, 'transaction_suggeree': None, 'statut': 'sans_correspondance'})

    transactions_non_matchees = [t for t in non_rapprochees if t['id'] not in ids_deja_suggeres]

    return jsonify({
        'nb_lignes_releve': len(lignes_releve),
        'nb_correspondances': len(ids_deja_suggeres),
        'resultats': resultats,
        'transactions_sans_correspondance': transactions_non_matchees,
    })


@bp.route('/rapprochement/valider/<t_id>', methods=['PUT'])
@require_auth
@require_role('admin', 'comptable')
def valider_rapprochement(t_id):
    row = db.execute("SELECT id FROM transactions WHERE id=? AND ecole_id=?", (t_id, g.user['ecole_id'])).fetchone()
    if not row:
        return jsonify({'error': 'Transaction introuvable'}), 404
    db.execute("UPDATE transactions SET rapproche=1, date_rapprochement=? WHERE id=? AND ecole_id=?",
               (datetime.now().strftime('%Y-%m-%d'), t_id, g.user['ecole_id']))
    db.commit()
    return jsonify({'success': True})


@bp.route('/rapprochement/annuler/<t_id>', methods=['PUT'])
@require_auth
@require_role('admin', 'comptable')
def annuler_rapprochement(t_id):
    db.execute("UPDATE transactions SET rapproche=0, date_rapprochement=NULL WHERE id=? AND ecole_id=?", (t_id, g.user['ecole_id']))
    db.commit()
    return jsonify({'success': True})


@bp.route('/rapprochement/etat', methods=['GET'])
@require_auth
@require_role('admin', 'directeur', 'comptable')
def etat_rapprochement():
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    sql = "SELECT * FROM transactions WHERE ecole_id=? AND statut_validation IN ('auto','valide')"
    params = [g.user['ecole_id']]
    if date_debut: sql += " AND date_op>=?"; params.append(date_debut)
    if date_fin: sql += " AND date_op<=?"; params.append(date_fin)
    sql += " ORDER BY date_op DESC"
    toutes = rows_to_list(db.execute(sql, params).fetchall())
    rapprochees = [t for t in toutes if t.get('rapproche')]
    return jsonify({
        'total': len(toutes), 'rapprochees': len(rapprochees),
        'non_rapprochees': len(toutes) - len(rapprochees),
        'transactions': toutes,
    })
